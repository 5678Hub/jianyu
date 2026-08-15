#!/usr/bin/env python3
"""
从 gb2762_2025.json 生成核对 Excel
3 个 sheet:
  Sheet1: 按食品类别查询(每条数据行)
  Sheet2: 食品大类统计(每个 L1 类别汇总)
  Sheet3: 脚注说明
"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

JSON_PATH = "data/gb2762/gb2762_2025.json"
OUT_PATH = "data/gb2762/gb2762_2025_食品类别.xlsx"

# 脚注收集
FOOTNOTES = {}  # 脚注字母 -> 完整文本

def get_foot(text):
    """提取一条数据行的脚注字母"""
    # 先看现有 JSON 的 remark/limits 字段
    return ""

def main():
    with open(JSON_PATH, 'r', encoding='utf-8') as f:
        data = json.load(f)

    contaminants = data['contaminants']

    # 收集脚注
    foot_text = ""
    for tab in contaminants:
        foot_text += tab.get("footnote", "") + "\n"

    # ========== Sheet 1: 按食品类别查询 ==========
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "按食品类别查询"

    headers1 = ["食品大类", "A.1 食品分类", "食品中文名称", "限量要求", "污染物", "元素符号", "单位", "脚注", "检验方法"]
    ws1.append(headers1)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin = Side(border_style="thin", color="B0B0B0")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for col in range(1, len(headers1) + 1):
        cell = ws1.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # 按 L1 (大类) → L2 (A.1 食品分类) → L3 (食品名) 顺序输出
    for tab in contaminants:
        cname = tab["contaminant"]
        symbol = tab.get("symbol", "")
        unit = tab["unit"]
        method = tab.get("inspection_method", "")

        for it in tab["items"]:
            cat1 = it.get("category", "")
            cat_a1 = it.get("category_a1", "")
            food = it.get("food", "")
            limit_val = it.get("limit", "")
            remark = it.get("remark", "")
            footnote = ""
            if remark:
                footnote = remark
            else:
                # 默认脚注:基于 GB 2762 表 1 铅: a 稻谷以糙米计。b 新鲜香辛料... c 液态婴幼儿...
                if tab["table_no"] == 1:
                    # 根据 cat_a1 推断（简化：a 全局标注）
                    pass

            # 拼接 limit_value + unit
            limit_str = f"{limit_val}{unit}"
            # 检查特殊 unit
            if "mg/L" in limit_val or "mg/L" in limit_str:
                limit_str = f"{limit_val}"
            ws1.append([
                cat1, cat_a1, food, limit_str,
                cname, symbol, unit, footnote, method
            ])

    # 列宽
    widths1 = [16, 30, 50, 12, 8, 8, 8, 60, 50]
    for i, w in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ========== Sheet 2: 食品大类统计 ==========
    ws2 = wb.create_sheet("食品大类统计")
    headers2 = ["食品大类", "数据条数", "覆盖污染物", "食品名示例"]
    ws2.append(headers2)
    for col in range(1, len(headers2) + 1):
        cell = ws2.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # 聚合
    cat_summary = {}  # L1_name -> { count, contaminants: set, foods: list }
    for tab in contaminants:
        for it in tab["items"]:
            cat1 = it.get("category", "(空)")
            if cat1 not in cat_summary:
                cat_summary[cat1] = {"count": 0, "contaminants": set(), "foods": []}
            cat_summary[cat1]["count"] += 1
            cat_summary[cat1]["contaminants"].add(tab["contaminant"])
            food = it.get("food", "")
            if food and food not in cat_summary[cat1]["foods"]:
                cat_summary[cat1]["foods"].append(food)

    # 排序输出
    for cat1 in sorted(cat_summary.keys(), key=lambda x: (x == "(空)", x)):
        s = cat_summary[cat1]
        ws2.append([
            cat1,
            s["count"],
            "、".join(sorted(s["contaminants"])),
            " / ".join(s["foods"][:3])
        ])

    widths2 = [16, 10, 35, 50]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ========== Sheet 3: 脚注说明 ==========
    ws3 = wb.create_sheet("脚注说明")
    headers3 = ["脚注", "所属污染物", "完整说明"]
    ws3.append(headers3)
    for col in range(1, len(headers3) + 1):
        cell = ws3.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # 标准脚注（来自 GB 2762-2025 表 1-12）
    standard_footnotes = [
        ("a", "铅、镉、汞、砷", "稻谷以糙米计。"),
        ("b", "铅(表1)", "新鲜香辛料(如姜、葱、蒜等)应按对应的新鲜蔬菜(或新鲜水果)类别执行。"),
        ("c", "铅(表1)", "液态婴幼儿配方食品根据 8:1 的比例折算其限量。"),
        ("b", "汞(表3)", "稻谷以糙米计。"),
        ("a", "汞(表3)", "对于制定甲基汞限量的食品可先测定其总汞含量,当总汞含量不超过甲基汞限量值时,可判定符合限量要求而不必测定甲基汞;否则,需测定甲基汞含量再作判定。"),
        ("a", "砷(表4)", "对于制定无机砷限量的食品可先测定其总砷含量,当总砷含量不超过无机砷限量值时,可判定符合限量要求而不必测定无机砷;否则,需测定无机砷含量再作判定。"),
        ("b", "砷(表4)", "稻谷以糙米计。"),
        ("a", "锡(表5)", "仅限于采用镀锡薄钢板容器包装的食品。"),
        ("a", "多氯联苯(表11)", "多氯联苯以 PCB28、PCB52、PCB101、PCB118、PCB138、PCB153 和 PCB180 的总和计。"),
        ("a", "3-氯-1,2-丙二醇(表12)", "仅限于添加酸水解植物蛋白的产品。"),
        ("a", "亚硝酸盐(表8)", "液态婴幼儿配方食品根据 8:1 的比例折算其限量。"),
        ("b", "亚硝酸盐(表8)", "仅适用于乳基产品。"),
        ("c", "亚硝酸盐(表8)", "液态婴幼儿配方食品根据 8:1 的比例折算其限量。"),
        ("d", "亚硝酸盐(表8)", "以固态产品计。"),
        ("e", "亚硝酸盐(表8)", "以固态产品计。"),
    ]

    for fn, ctxt, desc in standard_footnotes:
        ws3.append([fn, ctxt, desc])

    widths3 = [6, 22, 80]
    for i, w in enumerate(widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # 保存
    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    wb.save(OUT_PATH)
    print(f"✅ 已保存: {OUT_PATH}")

if __name__ == "__main__":
    main()
