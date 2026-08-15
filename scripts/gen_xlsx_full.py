#!/usr/bin/env python3
"""
生成最终版 Excel(完整版):
- Sheet 1: 按食品类别查询(12 列,含 4 级 A.1)
- Sheet 2: 食品大类统计
- Sheet 3: 脚注说明(完整版含所有表 1-12 脚注)
- Sheet 4: 按食品类别聚合(跨污染物,完整版)
  - 表 3 拆 总汞 + 甲基汞 a
  - 表 4 拆 总砷 + 无机砷 a
  - 表 8 拆 亚硝酸盐 + 硝酸盐
  - 污染物按 PDF 顺序排序
"""
import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from collections import defaultdict

JSON_PATH = "data/gb2762/gb2762_2025.json"
TMP_PATH = "data/gb2762/_gb2762_2025_食品类别_v5.xlsx"
OUT_PATH = "data/gb2762/gb2762_2025_食品类别.xlsx"


# 完整脚注(从 PDF 提取)
FOOTNOTES = [
    ("表 1 铅", "a", "稻谷以糙米计。"),
    ("表 1 铅", "b", "新鲜香辛料(如姜、葱、蒜等)应按对应的新鲜蔬菜(或新鲜水果)类别执行。"),
    ("表 1 铅", "c", "液态婴幼儿配方食品根据 8:1 的比例折算其限量。"),
    ("表 2 镉", "a", "稻谷以糙米计。"),
    ("表 3 汞", "a", "对于制定甲基汞限量的食品可先测定其总汞,当总汞含量不超过甲基汞限量值时,可判定符合限量要求而不必测定甲基汞;否则,需测定甲基汞含量再作判定。"),
    ("表 3 汞", "b", "稻谷以糙米计。"),
    ("表 4 砷", "a", "对于制定无机砷限量的食品可先测定其总砷,当总砷含量不超过无机砷限量值时,可判定符合限量要求而不必测定无机砷;否则,需测定无机砷含量再作判定。"),
    ("表 4 砷", "b", "稻谷以糙米计。"),
    ("表 5 锡", "a", "仅限于采用镀锡薄钢板容器包装的食品。"),
    ("表 8 亚硝酸盐", "a", "液态婴幼儿配方食品根据 8:1 的比例折算其限量。"),
    ("表 8 亚硝酸盐", "b", "仅适用于乳基产品。"),
    ("表 8 硝酸盐", "c", "不适用于添加蔬菜和水果的产品。"),
    ("表 8 硝酸盐", "d", "不适用于添加豆类的产品。"),
    ("表 8 硝酸盐", "e", "仅适用于乳基产品,不含豆类成分。"),
    ("表 9 苯并[a]芘", "a", "稻谷以糙米计。"),
    ("表 11 多氯联苯", "a", "PCB28、PCB52、PCB101、PCB118、PCB138、PCB153 和 PCB180 的总和计。"),
    ("表 12 3-氯-1,2-丙二醇", "a", "仅限于添加酸水解植物蛋白的产品。"),
]

# 污染物排序(按 PDF 表号顺序)
POLLUTANT_ORDER = {
    "铅": 1,
    "镉": 2,
    "总汞": 3,
    "甲基汞 a": 4,
    "总砷": 5,
    "无机砷 a": 6,
    "锡": 7,
    "镍": 8,
    "铬": 9,
    "亚硝酸盐": 10,
    "硝酸盐": 11,
    "苯并[a]芘": 12,
    "N-二甲基亚硝胺": 13,
    "多氯联苯": 14,
    "3-氯-1,2-丙二醇": 15,
}


def normalize(s):
    if not s:
        return ""
    return (
        s.replace("（", "(")
        .replace("）", ")")
        .replace("：", ":")
        .replace("，", ",")
        .replace("；", ";")
        .replace(" ", "")
        .replace("\u3000", "")
    )


def short_name(s):
    if not s:
        return ""
    n = normalize(s)
    for sep in ["(", "（"]:
        idx = n.find(sep)
        if idx > 0:
            return n[:idx]
    return n


def extract_keywords(name):
    if not name:
        return set()
    n = normalize(name)
    keywords = set()
    cleaned = n
    while "(" in cleaned:
        s = cleaned.find("(")
        e = cleaned.find(")", s)
        if e == -1:
            break
        inner = cleaned[s + 1 : e]
        for kw in re.split(r"[,、:]", inner):
            kw = kw.strip().rstrip("等")
            for prefix in ["例如", "如", "比如", "例如:", "如:"]:
                if kw.startswith(prefix):
                    kw = kw[len(prefix) :].strip()
            if 2 <= len(kw) <= 10:
                keywords.add(kw)
        cleaned = cleaned[:s] + cleaned[e + 1 :]
    keywords.add(short_name(name))
    for kw in re.split(r"[,、]", cleaned):
        kw = kw.strip().rstrip("等")
        if 2 <= len(kw) <= 10:
            keywords.add(kw)
    return keywords - {""}


def tree_name_variants(name):
    full = normalize(name)
    short = short_name(name)
    return ({full, short} | extract_keywords(name)) - {""}


def build_index(tree):
    by_name = {}
    for l1 in tree:
        l1n = normalize(l1["name"])
        by_name[l1n] = ("L1", l1, None)
        for l2 in l1.get("children", []):
            l2n = normalize(l2["name"])
            by_name[l2n] = ("L2", l2, l1)
            for l3 in l2.get("children", []):
                l3n = normalize(l3["name"])
                by_name[l3n] = ("L3", l3, l2)
                for l4 in l3.get("children", []):
                    l4n = normalize(l4["name"])
                    by_name[l4n] = ("L4", l4, l3)
    return by_name


def match_a1(item, by_name, tree):
    cat_a1 = item.get("category_a1", "") or ""
    food = item.get("food", "") or ""
    food_norm = normalize(food)
    cat_a1_norm = normalize(cat_a1)
    L1 = L2 = L3 = L4 = ""

    if cat_a1_norm and cat_a1_norm in by_name:
        level, node, parent = by_name[cat_a1_norm]
        if level == "L1":
            L1 = node["name"]
        elif level == "L2":
            L1 = parent["name"]
            L2 = node["name"]
    else:
        cat = item.get("category", "")
        cat_norm = normalize(cat)
        if cat_norm in by_name:
            level, node, parent = by_name[cat_norm]
            if level == "L1":
                L1 = node["name"]
            elif level == "L2":
                L1 = parent["name"]
                L2 = node["name"]

    if L1 and food:
        if food.endswith("除外") or food.endswith("等除外") or "、除外" in food:
            return L1, "", "", ""
        food_clean = food.rstrip("、,").rstrip()
        food_norm = normalize(food_clean)
        l1_node = None
        for n in tree:
            if n["name"] == L1:
                l1_node = n
                break
        if l1_node:
            candidates = []
            def walk(node, l2_name, l3_name, l4_name, depth):
                name = node.get("name", "")
                if name:
                    for variant in tree_name_variants(name):
                        if variant and variant in food_norm:
                            candidates.append((depth, l2_name, l3_name, l4_name, len(variant)))
                            break
                if depth >= 3:
                    return
                for child in node.get("children", []):
                    cname = child.get("name", "")
                    if depth == 0:
                        walk(child, cname, "", "", 1)
                    elif depth == 1:
                        walk(child, l2_name, cname, "", 2)
                    elif depth == 2:
                        walk(child, l2_name, l3_name, cname, 3)
            for c in l1_node.get("children", []):
                walk(c, c.get("name", ""), "", "", 1)
            if candidates:
                candidates.sort(key=lambda x: (-x[0], -x[4]))
                best = candidates[0]
                _, L2_n, L3_n, L4_n, _ = best
                L2 = L2_n if L2_n else L2
                L3 = L3_n
                L4 = L4_n
    return L1, L2, L3, L4


def apply_a1_to_items(data):
    """为每条 item 标记 a1_l1/a1_l2/a1_l3/a1_l4"""
    tree = data["appendix_a1"]["tree"]
    by_name = build_index(tree)
    for tab in data["contaminants"]:
        for it in tab["items"]:
            L1, L2, L3, L4 = match_a1(it, by_name, tree)
            it["a1_l1"] = L1
            it["a1_l2"] = L2
            it["a1_l3"] = L3
            it["a1_l4"] = L4


def gen_xlsx(data):
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="2F5496", end_color="2F5496", fill_type="solid"
    )
    thin = Side(border_style="thin", color="B0B0B0")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    def style_header(ws, cols):
        for col in range(1, cols + 1):
            c = ws.cell(1, col)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")

    # ========== Sheet 1: 按食品类别查询(12 列,含 4 级 A.1) ==========
    ws1 = wb.active
    ws1.title = "按食品类别查询"
    headers = [
        "食品大类",
        "A.1 食品分类(一级分类)",
        "A.1 食品分类(二级分类)",
        "A.1 食品分类(三级分类)",
        "A.1 食品分类(四级分类)",
        "食品中文名称",
        "限量要求",
        "污染物",
        "元素符号",
        "单位",
        "脚注",
        "检验方法",
    ]
    ws1.append(headers)
    style_header(ws1, 12)

    cur_L1 = cur_L2 = cur_L3 = cur_L4 = None
    for tab in data["contaminants"]:
        for it in tab["items"]:
            L1 = it.get("a1_l1", "") or ""
            L2 = it.get("a1_l2", "") or ""
            L3 = it.get("a1_l3", "") or ""
            L4 = it.get("a1_l4", "") or ""
            # 限量字符串
            limit_str = it.get("limit", "")
            unit = tab["unit"]
            if limit_str and limit_str != "—" and "mg/L" not in limit_str and "/L" not in limit_str:
                limit_str = f"{limit_str}{unit}"
            row = [
                it.get("category", ""),
                "" if L1 == cur_L1 else L1,
                "" if L2 == cur_L2 or (cur_L1 and L1 != cur_L1) else L2,
                "" if L3 == cur_L3 or (cur_L2 and L2 != cur_L2) else L3,
                "" if L4 == cur_L4 or (cur_L3 and L3 != cur_L3) else L4,
                it.get("food", ""),
                limit_str,
                tab["contaminant"],
                tab.get("symbol", ""),
                tab["unit"],
                it.get("remark", "") or "",
                tab.get("inspection_method", ""),
            ]
            ws1.append(row)
            cur_L1, cur_L2, cur_L3, cur_L4 = L1, L2, L3, L4

    widths = [16, 24, 24, 24, 24, 50, 16, 14, 10, 8, 50, 50]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[chr(64 + i)].width = w
    ws1.freeze_panes = "A2"

    # ========== Sheet 2: 食品大类统计 ==========
    ws2 = wb.create_sheet("食品大类统计")
    ws2.append(["食品大类", "数据条数", "覆盖污染物", "食品名示例"])
    style_header(ws2, 4)
    cat_sum = defaultdict(lambda: {"n": 0, "cn": set(), "foods": set()})
    for tab in data["contaminants"]:
        for it in tab["items"]:
            c1 = it.get("category", "(空)")
            cat_sum[c1]["n"] += 1
            cat_sum[c1]["cn"].add(tab["contaminant"])
            cat_sum[c1]["foods"].add(it.get("food", ""))
    for c1 in sorted(cat_sum.keys(), key=lambda x: (x == "(空)", x)):
        s = cat_sum[c1]
        ws2.append([c1, s["n"], "、".join(sorted(s["cn"])), " / ".join(sorted(s["foods"])[:3])])
    for col, w in zip("ABCD", [22, 12, 50, 60]):
        ws2.column_dimensions[col].width = w

    # ========== Sheet 3: 脚注说明(完整版) ==========
    ws3 = wb.create_sheet("脚注说明")
    ws3.append(["所属污染物表", "脚注", "完整说明"])
    style_header(ws3, 3)
    for tbl, fn, desc in FOOTNOTES:
        ws3.append([tbl, fn, desc])
    for col, w in zip("ABC", [22, 6, 80]):
        ws3.column_dimensions[col].width = w

    # ========== Sheet 4: 按食品类别聚合(跨污染物,完整版) ==========
    # 列: 食品大类 | A.1 L1 | A.1 L2 | A.1 L3 | A.1 L4 | 食品中文名称 | 污染物 | 限量值 | 单位 | 脚注 | 检验方法
    ws4 = wb.create_sheet("按食品类别聚合(完整版)")
    ws4.append(
        [
            "食品大类",
            "A.1 食品分类(一级分类)",
            "A.1 食品分类(二级分类)",
            "A.1 食品分类(三级分类)",
            "A.1 食品分类(四级分类)",
            "食品中文名称",
            "污染物",
            "限量值",
            "单位",
            "脚注",
            "检验方法",
        ]
    )
    style_header(ws4, 11)

    # 收集所有(食品, 污染物) 行
    rows = []
    for tab in data["contaminants"]:
        for it in tab["items"]:
            L1 = it.get("a1_l1", "") or ""
            L2 = it.get("a1_l2", "") or ""
            L3 = it.get("a1_l3", "") or ""
            L4 = it.get("a1_l4", "") or ""
            food = it.get("food", "")
            unit = tab["unit"]

            # 主污染物
            main_pollutant = tab["contaminant"]
            main_value = it.get("limit", "")
            main_remark = it.get("remark", "")

            # 子污染物(表 3/4/8)
            sub_pollutant = it.get("sub_label", "")
            sub_value = it.get("sub_limit", "")
            sub_remark = it.get("remark", "")

            # 行1: 主污染物
            main_limit_str = main_value
            if main_value and main_value != "—" and "mg/L" not in main_value and "/L" not in main_value:
                main_limit_str = f"{main_value}{unit}"
            rows.append(
                {
                    "category": it.get("category", ""),
                    "L1": L1, "L2": L2, "L3": L3, "L4": L4,
                    "food": food,
                    "pollutant": main_pollutant,
                    "value": main_limit_str,
                    "unit": unit,
                    "remark": main_remark,
                    "method": tab.get("inspection_method", ""),
                    "pollutant_order": POLLUTANT_ORDER.get(main_pollutant, 99),
                    "food_order": it.get("food", ""),
                }
            )
            # 行2: 子污染物(如适用)
            if sub_pollutant and sub_value:
                sub_limit_str = sub_value
                if sub_value != "—" and "mg/L" not in sub_value and "/L" not in sub_value:
                    sub_limit_str = f"{sub_value}{unit}"
                rows.append(
                    {
                        "category": it.get("category", ""),
                        "L1": L1, "L2": L2, "L3": L3, "L4": L4,
                        "food": food,
                        "pollutant": sub_pollutant,
                        "value": sub_limit_str,
                        "unit": unit,
                        "remark": sub_remark,
                        "method": tab.get("inspection_method", ""),
                        "pollutant_order": POLLUTANT_ORDER.get(sub_pollutant, 99),
                        "food_order": it.get("food", ""),
                    }
                )

    # 按 L1→L2→L3→L4→food→pollutant_order 排序
    rows.sort(
        key=lambda r: (
            r["L1"], r["L2"], r["L3"], r["L4"], r["food_order"], r["pollutant_order"]
        )
    )

    cur_L1 = cur_L2 = cur_L3 = cur_L4 = None
    for r in rows:
        ws4.append(
            [
                r["category"],
                "" if r["L1"] == cur_L1 else r["L1"],
                "" if r["L2"] == cur_L2 or (cur_L1 and r["L1"] != cur_L1) else r["L2"],
                "" if r["L3"] == cur_L3 or (cur_L2 and r["L2"] != cur_L2) else r["L3"],
                "" if r["L4"] == cur_L4 or (cur_L3 and r["L3"] != cur_L3) else r["L4"],
                r["food"],
                r["pollutant"],
                r["value"],
                r["unit"],
                r["remark"],
                r["method"],
            ]
        )
        cur_L1, cur_L2, cur_L3, cur_L4 = r["L1"], r["L2"], r["L3"], r["L4"]

    widths = [16, 24, 24, 24, 24, 50, 14, 16, 8, 50, 50]
    for i, w in enumerate(widths, 1):
        ws4.column_dimensions[chr(64 + i)].width = w
    ws4.freeze_panes = "A2"

    wb.save(TMP_PATH)
    print(f"✅ 已保存 {TMP_PATH}")
    print(f"Sheet 1: {ws1.max_row} 行(12 列)")
    print(f"Sheet 2: {ws2.max_row} 行")
    print(f"Sheet 3: {ws3.max_row} 行(完整脚注)")
    print(f"Sheet 4: {ws4.max_row} 行(按 L1→L2→L3→L4→食物→污染物 排序)")
    return TMP_PATH


def main():
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)
    apply_a1_to_items(data)
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print("已重新映射 A.1 4 级分类")
    gen_xlsx(data)


if __name__ == "__main__":
    main()