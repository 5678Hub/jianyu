import openpyxl
wb = openpyxl.load_workbook('data/gb2762/_gb2762_2025_食品类别_v6.xlsx', data_only=True)
ws = wb['按食品类别查询']

print("=== Sheet 1 表 1 铅(前 10 行) ===")
count = 0
for r in range(2, ws.max_row+1):
    if ws.cell(r, 8).value == '铅':
        cells = [str(ws.cell(r, c).value or '')[:18] for c in range(1, 13)]
        print('  | '.join(cells))
        count += 1
        if count >= 10:
            break

print(f"\n=== 表 1 铅 总数: {sum(1 for r in range(2, ws.max_row+1) if ws.cell(r, 8).value == '铅')} ===")
print(f"=== 表 2 镉 总数: {sum(1 for r in range(2, ws.max_row+1) if ws.cell(r, 8).value == '镉')} ===")
print(f"=== 表 3 汞 总数: {sum(1 for r in range(2, ws.max_row+1) if ws.cell(r, 8).value in ('总汞','甲基汞 a'))} ===")
print(f"=== 表 4 砷 总数: {sum(1 for r in range(2, ws.max_row+1) if ws.cell(r, 8).value in ('总砷','无机砷 a'))} ===")
print(f"=== 表 5 锡 总数: {sum(1 for r in range(2, ws.max_row+1) if ws.cell(r, 8).value == '锡')} ===")
print(f"=== 表 6 镍 总数: {sum(1 for r in range(2, ws.max_row+1) if ws.cell(r, 8).value == '镍')} ===")
print(f"=== 表 7 铬 总数: {sum(1 for r in range(2, ws.max_row+1) if ws.cell(r, 8).value == '铬')} ===")
print(f"=== 表 8 亚硝酸盐/硝酸盐 总数: {sum(1 for r in range(2, ws.max_row+1) if ws.cell(r, 8).value in ('亚硝酸盐','硝酸盐'))} ===")
print(f"=== 表 9 苯并[a]芘 总数: {sum(1 for r in range(2, ws.max_row+1) if ws.cell(r, 8).value == '苯并[a]芘')} ===")
print(f"=== 表 10 NDMA 总数: {sum(1 for r in range(2, ws.max_row+1) if ws.cell(r, 8).value == 'N-二甲基亚硝胺')} ===")
print(f"=== 表 11 多氯联苯 总数: {sum(1 for r in range(2, ws.max_row+1) if ws.cell(r, 8).value == '多氯联苯')} ===")
print(f"=== 表 12 3-氯丙二醇 总数: {sum(1 for r in range(2, ws.max_row+1) if ws.cell(r, 8).value == '3-氯-1,2-丙二醇')} ===")