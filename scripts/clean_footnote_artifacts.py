#!/usr/bin/env python3
"""
清理 JSON 数据中的 PDF 抓取脚注占位符 + 生成最终版 Excel
"""
import json
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict

JSON_PATH = "data/gb2762/gb2762_2025.json"
OUT_PATH = "data/gb2762/gb2762_2025_食品类别.xlsx"  # 最终版文件名


def clean_food(text):
    """清理 PDF 抓取脚注占位符"""
    if not text:
        return text
    s = text
    # 1. 开括号 + 2+ 顿号 → 整段删除(无论后面有没有右括号)
    s = re.sub(r'[(（][、,， ]{2,}', '', s)
    # 2. 2+ 顿号 + 右括号 → 整段删除
    s = re.sub(r'[、,， ]{2,}[)）]', '', s)
    # 3. 开括号 + 0~1 字符 + 右括号 → 整段删除(空括号、(、)、(,)、())
    s = re.sub(r'[(（][、,， ]{0,1}[)）]', '', s)
    # 4. 删除开头/结尾的连续顿号/逗号/空格
    s = re.sub(r'^[、,， ]+', '', s)
    s = re.sub(r'[、,， ]+$', '', s)
    # 5. 合并中间连续顿号(2+ 全部删)
    s = re.sub(r'[、,， ]{2,}', '', s)
    return s.strip()


def main():
    # 清理 JSON
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)

    cleaned_count = 0
    for tab in data["contaminants"]:
        for it in tab["items"]:
            old = it.get("food", "")
            new = clean_food(old)
            if old != new:
                cleaned_count += 1
                it["food"] = new
    print(f"已清理 {cleaned_count} 条 food 字段")

    # 写回 JSON(保持原 JSON 结构,只改 food)
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    # 删除所有旧 Excel
    for fname in os.listdir("data/gb2762"):
        if fname.endswith(".xlsx") and fname.startswith("gb2762_2025_食品类别"):
            os.remove(f"data/gb2762/{fname}")
            print(f"  删除: {fname}")

    # 收集 L1 出现顺序
    seen = []
    for tab in data["contaminants"]:
        for it in tab["items"]:
            c = it.get("category", "")
            if c and c not in seen:
                seen.append(c)
    L1_RANK = {c: i for i, c in enumerate(seen)}

    # 收集所有条目
    all_items = []
    for tab in data["contaminants"]:
        for it in tab["items"]:
            all_items.append({
                "L1": it.get("category", ""),
                "L2": it.get("category_a1", ""),
                "L3": it.get("food", ""),
                "污染物": tab["contaminant"],
                "元素符号": tab.get("symbol", ""),
                "限量": it.get("limit", ""),
                "单位": tab["unit"],
                "脚注": it.get("remark", "") or "",
                "检验方法": tab.get("inspection_method", ""),
            })

    all_items.sort(key=lambda it: (L1_RANK.get(it["L1"], 999), it["L1"], it["L2"], it["L3"], it["污染物"]))

    # 创建工作簿
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    def style_header(ws, cols):
        for col in range(1, cols + 1):
            c = ws.cell(1, col)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")

    # Sheet 1: 按食品类别查询(按污染物)
    ws1 = wb.active
    ws1.title = "按食品类别查询"
    ws1.append([
        "食品大类", "A.1 食品分类", "食品中文名称", "限量要求",
        "污染物", "元素符号", "单位", "脚注", "检验方法"
    ])
    style_header(ws1, 9)
    for tab in data["contaminants"]:
        for it in tab["items"]:
            limit = it.get("limit", "")
            unit = tab["unit"]
            limit_str = limit if "mg/L" in limit else f"{limit}{unit}"
            ws1.append([
                it.get("category", ""),
                it.get("category_a1", ""),
                it.get("food", ""),
                limit_str,
                tab["contaminant"],
                tab.get("symbol", ""),
                tab["unit"],
                it.get("remark", "") or "",
                tab.get("inspection_method", "")
            ])
    for col, w in zip("ABCDEFGHI", [18, 28, 50, 16, 10, 8, 8, 30, 50]):
        ws1.column_dimensions[col].width = w

    # Sheet 2: 食品大类统计
    ws2 = wb.create_sheet("食品大类统计")
    ws2.append(["食品大类", "数据条数", "覆盖污染物", "食品名示例"])
    style_header(ws2, 4)
    cat_sum = defaultdict(lambda: {"n": 0, "cn": set(), "foods": set()})
    for it in all_items:
        cat_sum[it["L1"]]["n"] += 1
        cat_sum[it["L1"]]["cn"].add(it["污染物"])
        cat_sum[it["L1"]]["foods"].add(it["L3"])
    for c1 in sorted(cat_sum.keys(), key=lambda x: L1_RANK.get(x, 999)):
        s = cat_sum[c1]
        ws2.append([
            c1, s["n"],
            "、".join(sorted(s["cn"])),
            " / ".join(sorted(s["foods"])[:3])
        ])
    for col, w in zip("ABCD", [22, 12, 50, 60]):
        ws2.column_dimensions[col].width = w

    # Sheet 3: 脚注说明
    ws3 = wb.create_sheet("脚注说明")
    ws3.append(["脚注", "所属污染物", "完整说明"])
    style_header(ws3, 3)
    fns = [
        ("a", "铅(表1)、镉(表2)", "稻谷以糙米计。"),
        ("b", "铅(表1)", "新鲜香辛料(如姜、葱、蒜等)应按对应的新鲜蔬菜(或新鲜水果)类别执行。"),
        ("c", "铅(表1)、亚硝酸盐(表8)", "液态婴幼儿配方食品根据 8:1 的比例折算其限量。"),
        ("a", "汞(表3)", "对于制定甲基汞限量的食品可先测定其总汞,当总汞含量不超过甲基汞限量值时,可判定符合限量要求而不必测定甲基汞;否则,需测定甲基汞含量再作判定。"),
        ("b", "汞(表3)", "稻谷以糙米计。"),
        ("a", "砷(表4)", "对于制定无机砷限量的食品可先测定其总砷,当总砷含量不超过无机砷限量值时,可判定符合限量要求而不必测定无机砷;否则,需测定无机砷含量再作判定。"),
        ("b", "砷(表4)", "稻谷以糙米计。"),
        ("a", "锡(表5)", "仅限于采用镀锡薄钢板容器包装的食品。"),
        ("a", "多氯联苯(表11)", "PCB28、PCB52、PCB101、PCB118、PCB138、PCB153 和 PCB180 的总和计。"),
        ("a", "3-氯-1,2-丙二醇(表12)", "仅限于添加酸水解植物蛋白的产品。"),
        ("b", "亚硝酸盐(表8)", "仅适用于乳基产品。"),
        ("d", "亚硝酸盐(表8)", "以固态产品计。"),
        ("e", "亚硝酸盐(表8)", "以固态产品计。"),
    ]
    for f, c1, d in fns:
        ws3.append([f, c1, d])
    for col, w in zip("ABC", [6, 22, 80]):
        ws3.column_dimensions[col].width = w

    # Sheet 4: 按食品类别聚合(跨污染物)
    ws4 = wb.create_sheet("按食品类别聚合(跨污染物)")
    ws4.append([
        "L1 大类", "L2 A.1 分类", "L3 食品名",
        "污染物", "元素符号", "限量", "单位", "检验方法"
    ])
    style_header(ws4, 8)
    cur_L1 = None
    cur_L2 = None
    for it in all_items:
        L1 = it["L1"]
        L2 = it["L2"]
        limit = it["限量"]
        unit = it["单位"]
        limit_str = limit if "mg/L" in limit else f"{limit}{unit}"
        l1_cell = L1 if L1 != cur_L1 else ""
        l2_cell = L2 if (L1 != cur_L1 or L2 != cur_L2) else ""
        cur_L1 = L1
        cur_L2 = L2
        ws4.append([
            l1_cell, l2_cell, it["L3"],
            it["污染物"], it["元素符号"],
            limit_str, unit, it["检验方法"]
        ])
    for col, w in zip("ABCDEFGH", [18, 28, 50, 14, 8, 12, 8, 50]):
        ws4.column_dimensions[col].width = w
    ws4.freeze_panes = "A2"

    wb.save(OUT_PATH)
    print(f"✅ 最终版已生成: {OUT_PATH}")
    print(f"  Sheet 1: {ws1.max_row} 行")
    print(f"  Sheet 2: {ws2.max_row} 行")
    print(f"  Sheet 3: {ws3.max_row} 行")
    print(f"  Sheet 4: {ws4.max_row} 行")


if __name__ == "__main__":
    main()
