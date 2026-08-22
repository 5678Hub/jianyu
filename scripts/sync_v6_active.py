"""
从当前活动 v6_synced.xlsx 同步到 JSON 和 jianyu-offline.html。

特点：
1. 同步所有 12 张污染物表（README/汇总统计/新增分类清单 不动）
2. note 列智能解析：提取 abc 字母；全文本时字母+文本分离
3. 构建/合并 footnotes 数组：
   - Excel 全文本（如 'a 稻谷以糙米计。'）→ label='a', text=后半部分
   - Excel 仅字母 → 复用已有 JSON 的 footnotes（保留旧文本）
   - 新增字母 → 添加新条目

用法：
  python scripts/sync_v6_active.py
"""
import argparse
import datetime
import json
import re
import shutil
from copy import deepcopy
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = ROOT / "data" / "gb2762" / "_gb2762_2025_12项污染物_v6_synced.xlsx"
JSON_PATH = ROOT / "data" / "gb2762" / "gb2762_2025.json"
OFFLINE = ROOT / "jianyu-offline.html"
META = ROOT / "data" / "gb2762" / "_meta"
DATA_START_ROW = 5

SKIP_SHEETS = {"README", "汇总统计", "🆕新增分类清单"}

# 12 个污染物表的元信息：(xlsx 表名, table_no, JSON key, 污染物名称, 结构类型)
TABLE_DEFS = [
    ("1.铅", 1, "铅", "simple"),
    ("2.镉", 2, "镉", "simple"),
    ("3.汞", 3, "汞", "simple"),
    ("4.砷", 4, "砷", "pair"),
    ("5.锡", 5, "锡", "category"),
    ("6.镍", 6, "镍", "category"),
    ("7.铬", 7, "铬", "main_only"),
    ("8.亚硝酸盐", 8, "亚硝酸盐、硝酸盐", "pair"),
    ("9.苯并a芘", 9, "苯并[a]芘", "main_only"),
    ("10.N二甲基亚硝胺", 10, "N-二甲基亚硝胺", "category"),
    ("11.多氯联苯", 11, "多氯联苯", "category"),
    ("12.氯丙二醇", 12, "3-氯-1,2-丙二醇", "category"),
]

# 解析 note：'<letter> <text>' → ('<letter>', '<text>')；仅字母 → ('<letter>', None)
NOTE_LETTER_RE = re.compile(r"^([a-zA-Z])\s*(.*)$")


def parse_note(raw_note):
    """解析脚注：返回 (letter, text) 元组；如果空则返回 (None, None)。"""
    if not raw_note:
        return None, None
    raw = str(raw_note).strip()
    if not raw:
        return None, None
    m = NOTE_LETTER_RE.match(raw)
    if m:
        letter = m.group(1)
        text = m.group(2).strip()
        return letter, text
    return None, raw  # 非字母开头：保留全文作为 text，letter 为空


def read_xlsx_rows(source, sheet_name):
    wb = load_workbook(source, data_only=True)
    ws = wb[sheet_name]
    rows = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        l1 = ws.cell(r, 1).value
        food = ws.cell(r, 5).value
        if not (l1 or food):
            continue
        rows.append({
            "a1_l1": l1 or "",
            "a1_l2": ws.cell(r, 2).value or "",
            "a1_l3": ws.cell(r, 3).value or "",
            "a1_l4": ws.cell(r, 4).value or "",
            "food": food or "",
            "pollutant": ws.cell(r, 6).value or "",
            "limit_value": ws.cell(r, 7).value or "",
            "modif": ws.cell(r, 8).value or "",
            "unit": ws.cell(r, 9).value or "",
            "raw_note": ws.cell(r, 10).value or "",
            "inspection_method": ws.cell(r, 11).value or "",
        })
    return rows


def collect_footnotes(rows, existing_footnotes):
    """从行 + 已有 footnotes 中合并构建 footnotes 列表。

    规则：
    1. 遍历 rows 提取 (letter, text)；
    2. text 不为空时优先用 Excel 的 text；
    3. text 为空时复用已有 footnotes 中同 letter 的 text。
    4. 只返回 Excel 中实际引用的字母。
    5. 返回按字母排序的 list of {label, text}。
    """
    by_label = {}
    for fn in existing_footnotes or []:
        if isinstance(fn, dict) and fn.get("label"):
            by_label[fn["label"]] = fn.get("text", "")

    referenced = set()
    for r in rows:
        letter, text = parse_note(r["raw_note"])
        if not letter:
            continue
        referenced.add(letter)
        if text:
            by_label[letter] = text
        # else: 保留 by_label 中已存在的

    return [{"label": k, "text": by_label[k]} for k in sorted(by_label.keys()) if k in referenced]


def to_simple_json_item(row):
    letter, _ = parse_note(row["raw_note"])
    limit_str = str(row["limit_value"]).strip()
    has_limit = limit_str not in ("", "—", "-")
    return {
        "food": row["food"],
        "pollutant": row["pollutant"],
        "limit_value": limit_str,
        "has_limit": has_limit,
        "sub_value": "",
        "unit": row["unit"],
        "note": letter or "",  # 只存字母
        "modif": row["modif"],
        "inspection_method": row["inspection_method"],
        "a1_l1": row["a1_l1"],
        "a1_l2": row["a1_l2"],
        "a1_l3": row["a1_l3"],
        "a1_l4": row["a1_l4"],
    }


def to_category_json_item(row):
    """表 5/6/8/10/11/12: category-based 结构

    fields: category, food, limit, limit_value, a1_l1..4, remark (footnote letter), inspection_method, test_method
    """
    l2 = row["a1_l2"]
    l1 = row["a1_l1"]
    l3 = row["a1_l3"]
    category = l2 if l2 else l1

    letter, _ = parse_note(row["raw_note"])
    limit_str = f'{row["limit_value"]} {row["unit"]}'.strip()
    limit_val = str(row["limit_value"]).strip()
    has_limit = limit_val not in ("", "—", "-")

    item = {
        "category": category or "",
        "food": row["food"],
        "limit": limit_str,
        "limit_value": limit_val,
        "has_limit": has_limit,
        "a1_l1": row["a1_l1"],
        "a1_l2": row["a1_l2"],
        "a1_l3": row["a1_l3"],
        "a1_l4": row["a1_l4"],
        "remark": letter or "",
        "note": letter or "",
        "inspection_method": row["inspection_method"],
        "test_method": "",
    }
    return item


def to_main_only_json_item(row, contaminant_name):
    letter, _ = parse_note(row["raw_note"])
    limit_val = str(row["limit_value"]).strip()
    has_limit = limit_val not in ("", "—", "-")
    return {
        "food": row["food"],
        "limit": f'{row["limit_value"]} {row["unit"]}'.strip(),
        "limit_value": limit_val,
        "has_limit": has_limit,
        "limit_modifier": row["modif"],
        "main_label": contaminant_name,
        "main_remark": "",  # 修饰不在 main_remark
        "sub_label": "",
        "sub_limit": "",
        "sub_value": "",
        "sub_modifier": "",
        "sub_remark": "",
        "remark": letter or "",  # 仅字母
        "a1_l1": row["a1_l1"],
        "a1_l2": row["a1_l2"],
        "a1_l3": row["a1_l3"],
        "a1_l4": row["a1_l4"],
        "inspection_method": row["inspection_method"],
        "test_method": "",
    }


def to_pair_json_item(rows_pair, contaminant_name):
    """rows_pair: 同一 food 的 2 行（总砷+无机砷 或 亚硝酸盐+硝酸盐）。
    识别规则：
    - 表 4 砷：含"无机"→ sub_row，其他→ main_row
    - 表 8 亚硝酸盐：含"硝酸盐"（但不是"亚硝酸盐"）→ sub_row
    - 其他：fallback 到第一行 main + 第二行 sub
    """
    main_row = None
    sub_row = None
    has_explicit_pair = False
    for r in rows_pair:
        pol = r["pollutant"] or ""
        if "无机" in pol:  # 表 4 砷：总砷 + 无机砷
            sub_row = r
            has_explicit_pair = True
        elif "亚硝酸盐" in pol:  # 表 8 亚硝酸盐：亚硝酸盐 + 硝酸盐
            main_row = r
            has_explicit_pair = True
        elif "硝酸盐" in pol:  # 表 8 亚硝酸盐：亚硝酸盐 + 硝酸盐
            sub_row = r
            has_explicit_pair = True
        else:
            # 没明确标识，按行号顺序：第一行 main，第二行 sub
            if main_row is None:
                main_row = r
            elif sub_row is None:
                sub_row = r

    if not has_explicit_pair:
        # fallback: 按行号顺序
        if main_row is None:
            main_row = rows_pair[0]
        if sub_row is None:
            sub_row = rows_pair[-1] if rows_pair[-1] is not main_row else (rows_pair[1] if len(rows_pair) > 1 else None)

    main_letter, _ = parse_note(main_row["raw_note"])
    sub_letter, _ = parse_note(sub_row["raw_note"])

    def fmt(row):
        return f'{row["limit_value"]} {row["unit"]}'.strip() if row else ""

    def has_limit(row):
        if not row:
            return False
        v = str(row["limit_value"]).strip()
        return v not in ("", "—", "-")

    return {
        "food": main_row["food"],
        "limit": fmt(main_row),
        "limit_value": str(main_row["limit_value"]),
        "has_limit": has_limit(main_row),
        "limit_modifier": main_row["modif"],
        "main_label": main_row["pollutant"] or "总砷",
        "main_remark": main_letter or "",
        "sub_label": sub_row["pollutant"] if sub_row else "无机砷 a",
        "sub_limit": fmt(sub_row) if sub_row else "—",
        "sub_value": str(sub_row["limit_value"]) if sub_row else "—",
        "sub_has_limit": has_limit(sub_row),
        "sub_modifier": sub_row["modif"] if sub_row else "",
        "sub_remark": sub_letter or "",
        "remark": main_letter or "",
        "a1_l1": main_row["a1_l1"],
        "a1_l2": main_row["a1_l2"],
        "a1_l3": main_row["a1_l3"],
        "a1_l4": main_row["a1_l4"],
        "inspection_method": main_row["inspection_method"],
        "test_method": "",
    }


def group_pair_rows(rows):
    """把表 4 / 表 8 的成对行按 (a1_l1, a1_l2, a1_l3, food) 分组。
    包含 a1_l3 是因为表 8 中部分行 L3 非空（如酱腌菜），不同 L3 相同 food 应分到不同组。
    """
    groups = {}
    for r in rows:
        key = (r["a1_l1"], r["a1_l2"], r["a1_l3"], r["food"])
        groups.setdefault(key, []).append(r)
    return list(groups.values())


def to_offline_item(row):
    l2 = row["a1_l2"]
    l1 = row["a1_l1"]
    l3 = row["a1_l3"]
    category = l2 if l2 else l1
    if l3:
        cat_matched = l3
    elif l2:
        cat_matched = l2
    else:
        cat_matched = row["food"][:10]

    letter, _ = parse_note(row["raw_note"])

    item = {
        "category": category,
        "food": row["food"],
        "limit": str(row["limit_value"]),
        "remark": row["modif"],
        "category_matched_by": cat_matched,
        "subcategories": [],
        "category_a1": l1,
        "limits": [{
            "label": f'限量(以 {row["pollutant"] or "Pb"} 计)',
            "value": str(row["limit_value"])
        }],
    }
    if l2:
        item["a1_l2"] = l2
    if l3:
        item["a1_l3"] = l3
    if row["a1_l4"]:
        item["a1_l4"] = row["a1_l4"]
    if row["unit"]:
        item["unit"] = row["unit"]
    if letter:
        item["note"] = letter
    return item


def backup_files():
    META.mkdir(exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    json_bak = META / f"gb2762_2025_pre_active_sync_{ts}.json"
    offline_bak = META / f"jianyu-offline_pre_active_sync_{ts}.html"
    shutil.copy(JSON_PATH, json_bak)
    shutil.copy(OFFLINE, offline_bak)
    print(f"[备份] {json_bak.name}")
    print(f"[备份] {offline_bak.name}")
    return ts


def sync_json(source):
    data = json.load(open(JSON_PATH, encoding="utf-8"))
    contam_map = {c.get("table_no"): c for c in data["contaminants"]}

    for sheet_name, table_no, contam_name, structure in TABLE_DEFS:
        if structure == "skip":
            print(f"  [跳过] 表 {table_no} {contam_name} (结构特殊)")
            continue
        rows = read_xlsx_rows(source, sheet_name)

        if structure == "simple":
            items = [to_simple_json_item(r) for r in rows]
        elif structure == "main_only":
            items = [to_main_only_json_item(r, contam_name) for r in rows]
        elif structure == "pair":
            grouped = group_pair_rows(rows)
            items = [to_pair_json_item(g, contam_name) for g in grouped]
        elif structure == "category":
            items = [to_category_json_item(r) for r in rows]

        contam = contam_map.get(table_no)
        if not contam:
            print(f"  [跳过] 表 {table_no} 不在 JSON 中")
            continue

        old_count = len(contam["items"])
        contam["items"] = items

        # 合并 footnotes
        old_footnotes = contam.get("footnotes", [])
        new_footnotes = collect_footnotes(rows, old_footnotes)
        contam["footnotes"] = new_footnotes

        print(f"  表 {table_no} {contam_name} ({structure}): {old_count} → {len(items)} items, footnotes: {[f['label'] for f in new_footnotes]}")

    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"[保存] {JSON_PATH.name}")


def sync_offline(source):
    rows = read_xlsx_rows(source, "1.铅")
    new_items = [to_offline_item(r) for r in rows]
    print(f"  v6 表 1: {len(rows)} 行 → {len(new_items)} items")

    content = open(OFFLINE, encoding="utf-8").read()
    m = re.search(r',"gb2762":\{', content)
    if not m:
        raise Exception("未找到 gb2762 数据")
    gb_s = m.end() - 1
    depth = 0
    gb_e = gb_s
    for i in range(gb_s, len(content)):
        ch = content[i]
        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                gb_e = i + 1
                break

    gb_dict = json.loads(content[gb_s:gb_e])
    t1_index = -1
    for i, c in enumerate(gb_dict["contaminants"]):
        if c.get("table_no") == 1:
            t1_index = i
            break
    if t1_index < 0:
        raise Exception("未找到 table_no=1")

    old = len(gb_dict["contaminants"][t1_index]["items"])
    gb_dict["contaminants"][t1_index]["items"] = new_items

    new_gb = json.dumps(gb_dict, ensure_ascii=False, separators=(",", ":"))
    new_content = content[:gb_s] + new_gb + content[gb_e:]
    with open(OFFLINE, "w", encoding="utf-8") as f:
        f.write(new_content)
    print(f"  [保存] {OFFLINE.name} (表 1: {old} → {len(new_items)} items)")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", default=str(DEFAULT_SOURCE),
                        help="源 xlsx 文件路径")
    parser.add_argument("--skip-backup", action="store_true")
    args = parser.parse_args()
    source = Path(args.source)
    if not source.exists():
        raise SystemExit(f"源文件不存在: {source}")

    print("=" * 60)
    print(f"活动 v6 → JSON + jianyu-offline.html 同步")
    print(f"源: {source.name}")
    print("=" * 60)

    if not args.skip_backup:
        backup_files()
    else:
        print("[跳过备份]")

    print()
    print("=== 同步 JSON ===")
    sync_json(source)

    print()
    print("=== 同步 jianyu-offline.html ===")
    sync_offline(source)

    print()
    print("=" * 60)
    print("同步完成")
    print("=" * 60)


if __name__ == "__main__":
    main()