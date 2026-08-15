#!/usr/bin/env python3
"""
为每条 item 添加 A.1 4 级分类(L1/L2/L3/L4),重新生成最终版 Excel(12 列)
"""
import json
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from collections import defaultdict

JSON_PATH = "data/gb2762/gb2762_2025.json"
TMP_PATH = "data/gb2762/_gb2762_2025_食品类别_v4.xlsx"
OUT_PATH = "data/gb2762/gb2762_2025_食品类别.xlsx"


def normalize(s):
    """统一字符,便于匹配"""
    if not s:
        return ""
    return (
        s.replace("（", "(")
        .replace("）", ")")
        .replace("：", ":")
        .replace("，", ",")
        .replace("；", ";")
        .replace(" ", "")
        .replace("\u3000", "")
    )


def short_name(s):
    """提取短名:取括号前的部分"""
    if not s:
        return ""
    n = normalize(s)
    for sep in ["(", "（"]:
        idx = n.find(sep)
        if idx > 0:
            return n[:idx]
    return n


def tree_name_variants(name):
    """生成 name 的多个匹配变体:全名 + 短名 + 关键词"""
    full = normalize(name)
    short = short_name(name)
    keywords = extract_keywords(name)
    return ({full, short} | keywords) - {""}


def extract_keywords(name):
    """从 tree name 提取关键词(去前缀'例如'等)"""
    if not name:
        return set()
    n = normalize(name)
    keywords = set()
    # 去掉括号内容(保留内部关键词)
    cleaned = n
    while "(" in cleaned:
        s = cleaned.find("(")
        e = cleaned.find(")", s)
        if e == -1:
            break
        # 提取括号内关键词(按 , 或 、 分隔)
        inner = cleaned[s + 1 : e]
        for kw in re.split(r"[,、:]", inner):
            kw = kw.strip().rstrip("等")
            # 去掉"例如"等前缀
            for prefix in ["例如", "如", "比如", "例如:", "如:"]:
                if kw.startswith(prefix):
                    kw = kw[len(prefix) :]
                    kw = kw.strip()
            if 2 <= len(kw) <= 10:
                keywords.add(kw)
        cleaned = cleaned[:s] + cleaned[e + 1 :]
    # 整个短名作为关键词
    keywords.add(short_name(name))
    # 从 cleaned 提取主关键词(去除"等"等)
    for kw in re.split(r"[,、]", cleaned):
        kw = kw.strip().rstrip("等")
        if 2 <= len(kw) <= 10:
            keywords.add(kw)
    return keywords - {""}


def build_index(tree):
    """建立 name -> node 索引 + L2 索引(用于 '水果制品' 这种 L2 当 L1 的情况)"""
    by_name = {}  # normalized name -> (level, node, parent_node)
    for l1 in tree:
        l1n = normalize(l1["name"])
        by_name[l1n] = ("L1", l1, None)
        for l2 in l1.get("children", []):
            l2n = normalize(l2["name"])
            by_name[l2n] = ("L2", l2, l1)
            for l3 in l2.get("children", []):
                l3n = normalize(l3["name"])
                by_name[l3n] = ("L3", l3, l2)
                for l4 in l3.get("children", []):
                    l4n = normalize(l4["name"])
                    by_name[l4n] = ("L4", l4, l3)
    return by_name


def match_a1(item, by_name, tree):
    """返回 (L1, L2, L3, L4)"""
    cat_a1 = item.get("category_a1", "") or ""
    food = item.get("food", "") or ""
    food_norm = normalize(food)
    cat_a1_norm = normalize(cat_a1)

    # 默认空
    L1 = L2 = L3 = L4 = ""

    # 1. 优先匹配 category_a1
    if cat_a1_norm and cat_a1_norm in by_name:
        level, node, parent = by_name[cat_a1_norm]
        if level == "L1":
            L1 = node["name"]
        elif level == "L2":
            L1 = parent["name"]
            L2 = node["name"]
        elif level == "L3":
            # L3 当 L1,找 L2/L3
            l2_node = parent
            l1_node = None
            for n in tree:
                if l2_node in n.get("children", []):
                    l1_node = n
                    break
            if l1_node:
                L1 = l1_node["name"]
                L2 = l2_node["name"]
                L3 = node["name"]
    else:
        # 没匹配上 category_a1,尝试用 category(L1 大类)
        cat = item.get("category", "")
        cat_norm = normalize(cat)
        if cat_norm in by_name:
            level, node, parent = by_name[cat_norm]
            if level == "L1":
                L1 = node["name"]
            elif level == "L2":
                L1 = parent["name"]
                L2 = node["name"]

    # 2. 在 L1 节点下找 L2/L3/L4(深度优先,匹配最长关键词)
    if L1 and food:
        # 如果 food 以"除外"结尾(整类条目),只到 L1
        if food.endswith("除外") or food.endswith("等除外") or "、除外" in food:
            return L1, "", "", ""
        # 去掉末尾顿号
        food_clean = food.rstrip("、,").rstrip()
        food_norm = normalize(food_clean)

        l1_node = None
        for n in tree:
            if n["name"] == L1:
                l1_node = n
                break
        if l1_node:
            # 收集所有候选匹配(L2/L3/L4 路径)
            candidates = []  # (depth, l2_name, l3_name, l4_name, match_len)

            def walk(node, l2_name, l3_name, l4_name, depth):
                name = node.get("name", "")
                if name:
                    # 尝试全名和短名匹配
                    for variant in tree_name_variants(name):
                        if variant and variant in food_norm:
                            candidates.append((depth, l2_name, l3_name, l4_name, len(variant)))
                            break  # 同一节点不重复添加
                # 限制递归深度 ≤ 3
                if depth >= 3:
                    return
                for child in node.get("children", []):
                    cname = child.get("name", "")
                    if depth == 0:
                        walk(child, cname, "", "", 1)
                    elif depth == 1:
                        walk(child, l2_name, cname, "", 2)
                    elif depth == 2:
                        walk(child, l2_name, l3_name, cname, 3)

            for c in l1_node.get("children", []):
                walk(c, c.get("name", ""), "", "", 1)

            # 选择最深的匹配(depth 最大),同 depth 时选最长 match
            if candidates:
                candidates.sort(key=lambda x: (-x[0], -x[4]))
                best = candidates[0]
                _, L2_n, L3_n, L4_n, _ = best
                L2 = L2_n if L2_n else L2
                L3 = L3_n
                L4 = L4_n

    return L1, L2, L3, L4


def main():
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)

    tree = data["appendix_a1"]["tree"]
    by_name = build_index(tree)

    # 为每条 item 添加 L2/L3/L4
    n_updated = 0
    for tab in data["contaminants"]:
        for it in tab["items"]:
            L1 = it.get("category_a1", "") or it.get("category", "")
            L2, L3, L4 = "", "", ""
            # 重新匹配以获得 4 级
            new_L1, new_L2, new_L3, new_L4 = match_a1(it, by_name, tree)
            if new_L1:
                L1 = new_L1
            L2 = new_L2
            L3 = new_L3
            L4 = new_L4
            it["a1_l1"] = L1
            it["a1_l2"] = L2
            it["a1_l3"] = L3
            it["a1_l4"] = L4
            n_updated += 1
    print(f"已为 {n_updated} 条 item 添加 A.1 4 级分类")

    # 写回 JSON
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"已写回 {JSON_PATH}")

    # === 生成 Excel ===
    generate_xlsx(data)


def generate_xlsx(data):
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="2F5496", end_color="2F5496", fill_type="solid"
    )
    L1_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    L2_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin = Side(border_style="thin", color="B0B0B0")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    def style_header(ws, cols):
        for col in range(1, cols + 1):
            c = ws.cell(1, col)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")

    # ===== Sheet 1: 按食品类别查询(12 列) =====
    ws1 = wb.active
    ws1.title = "按食品类别查询"
    headers = [
        "食品大类",
        "A.1 食品分类(一级分类)",
        "A.1 食品分类(二级分类)",
        "A.1 食品分类(三级分类)",
        "A.1 食品分类(四级分类)",
        "食品中文名称",
        "限量要求",
        "污染物",
        "元素符号",
        "单位",
        "脚注",
        "检验方法",
    ]
    ws1.append(headers)
    style_header(ws1, 12)

    cur_L1 = cur_L2 = cur_L3 = cur_L4 = None
    for tab_i, tab in enumerate(data["contaminants"]):
        for it in tab["items"]:
            L1 = it.get("a1_l1", "") or ""
            L2 = it.get("a1_l2", "") or ""
            L3 = it.get("a1_l3", "") or ""
            L4 = it.get("a1_l4", "") or ""
            limit_str = f"{it.get('limit','')}{tab['unit']}"
            if "mg/L" in it.get("limit", "") or "/L" in it.get("limit", ""):
                limit_str = it.get("limit", "")
            # 合并显示:相同 L1/L2/L3/L4 留空以视觉层级
            row = [
                it.get("category", ""),
                "" if L1 == cur_L1 else L1,
                "" if L2 == cur_L2 or (cur_L1 and L1 != cur_L1) else L2,
                "" if L3 == cur_L3 or (cur_L2 and L2 != cur_L2) else L3,
                "" if L4 == cur_L4 or (cur_L3 and L3 != cur_L3) else L4,
                it.get("food", ""),
                limit_str,
                tab["contaminant"],
                tab.get("symbol", ""),
                tab["unit"],
                it.get("remark", "") or "",
                tab.get("inspection_method", ""),
            ]
            ws1.append(row)
            cur_L1, cur_L2, cur_L3, cur_L4 = L1, L2, L3, L4

    widths = [16, 24, 24, 24, 24, 50, 16, 10, 8, 8, 50, 50]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[chr(64 + i)].width = w
    ws1.freeze_panes = "A2"

    # ===== Sheet 2: 食品大类统计 =====
    ws2 = wb.create_sheet("食品大类统计")
    ws2.append(["食品大类", "数据条数", "覆盖污染物", "食品名示例"])
    style_header(ws2, 4)
    cat_sum = defaultdict(lambda: {"n": 0, "cn": set(), "foods": set()})
    for tab in data["contaminants"]:
        for it in tab["items"]:
            c1 = it.get("category", "(空)")
            cat_sum[c1]["n"] += 1
            cat_sum[c1]["cn"].add(tab["contaminant"])
            cat_sum[c1]["foods"].add(it.get("food", ""))
    for c1 in sorted(cat_sum.keys(), key=lambda x: (x == "(空)", x)):
        s = cat_sum[c1]
        ws2.append([c1, s["n"], "、".join(sorted(s["cn"])), " / ".join(sorted(s["foods"])[:3])])
    for col, w in zip("ABCD", [22, 12, 50, 60]):
        ws2.column_dimensions[col].width = w

    # ===== Sheet 3: 脚注说明 =====
    ws3 = wb.create_sheet("脚注说明")
    ws3.append(["脚注", "所属污染物", "完整说明"])
    style_header(ws3, 3)
    fns = [
        ("a", "铅(表1)、镉(表2)", "稻谷以糙米计。"),
        ("b", "铅(表1)", "新鲜香辛料(如姜、葱、蒜等)应按对应的新鲜蔬菜(或新鲜水果)类别执行。"),
        ("c", "铅(表1)、亚硝酸盐(表8)", "液态婴幼儿配方食品根据 8:1 的比例折算其限量。"),
        ("a", "汞(表3)", "对于制定甲基汞限量的食品可先测定其总汞,当总汞含量不超过甲基汞限量值时,可判定符合限量要求而不必测定甲基汞;否则,需测定甲基汞含量再作判定。"),
        ("b", "汞(表3)", "稻谷以糙米计。"),
        ("a", "砷(表4)", "对于制定无机砷限量的食品可先测定其总砷,当总砷含量不超过无机砷限量值时,可判定符合限量要求而不必测定无机砷;否则,需测定无机砷含量再作判定。"),
        ("b", "砷(表4)", "稻谷以糙米计。"),
        ("a", "锡(表5)", "仅限于采用镀锡薄钢板容器包装的食品。"),
        ("a", "多氯联苯(表11)", "PCB28、PCB52、PCB101、PCB118、PCB138、PCB153 和 PCB180 的总和计。"),
        ("a", "3-氯-1,2-丙二醇(表12)", "仅限于添加酸水解植物蛋白的产品。"),
        ("b", "亚硝酸盐(表8)", "仅适用于乳基产品。"),
        ("d", "亚硝酸盐(表8)", "以固态产品计。"),
        ("e", "亚硝酸盐(表8)", "以固态产品计。"),
    ]
    for f, c1, d in fns:
        ws3.append([f, c1, d])
    for col, w in zip("ABC", [6, 22, 80]):
        ws3.column_dimensions[col].width = w

    # ===== Sheet 4: 按食品类别聚合(跨污染物) =====
    ws4 = wb.create_sheet("按食品类别聚合(跨污染物)")
    ws4.append(
        [
            "L1 A.1 一级分类",
            "L2 A.1 二级分类",
            "L3 A.1 三级分类",
            "L4 A.1 四级分类",
            "食品中文名称",
            "污染物",
            "元素符号",
            "限量",
            "单位",
            "检验方法",
        ]
    )
    style_header(ws4, 10)

    # 按 L1→L2→L3→L4→污染物 排序
    rows = []
    for tab in data["contaminants"]:
        for it in tab["items"]:
            rows.append(
                {
                    "L1": it.get("a1_l1", "") or it.get("category_a1", "") or "",
                    "L2": it.get("a1_l2", ""),
                    "L3": it.get("a1_l3", ""),
                    "L4": it.get("a1_l4", ""),
                    "食品": it.get("food", ""),
                    "污染物": tab["contaminant"],
                    "元素": tab.get("symbol", ""),
                    "限量": it.get("limit", ""),
                    "单位": tab["unit"],
                    "方法": tab.get("inspection_method", ""),
                }
            )
    rows.sort(key=lambda r: (r["L1"], r["L2"], r["L3"], r["L4"], r["食品"], r["污染物"]))

    cur_L1 = cur_L2 = cur_L3 = cur_L4 = None
    for r in rows:
        limit_str = f"{r['限量']}{r['单位']}" if "mg/L" not in r["限量"] and "/L" not in r["限量"] else r["限量"]
        ws4.append(
            [
                "" if r["L1"] == cur_L1 else r["L1"],
                "" if r["L2"] == cur_L2 or (cur_L1 and r["L1"] != cur_L1) else r["L2"],
                "" if r["L3"] == cur_L3 or (cur_L2 and r["L2"] != cur_L2) else r["L3"],
                "" if r["L4"] == cur_L4 or (cur_L3 and r["L3"] != cur_L3) else r["L4"],
                r["食品"],
                r["污染物"],
                r["元素"],
                limit_str,
                r["单位"],
                r["方法"],
            ]
        )
        cur_L1, cur_L2, cur_L3, cur_L4 = r["L1"], r["L2"], r["L3"], r["L4"]

    widths = [22, 24, 24, 24, 50, 14, 8, 12, 8, 50]
    for i, w in enumerate(widths, 1):
        ws4.column_dimensions[chr(64 + i)].width = w
    ws4.freeze_panes = "A2"

    # 保存
    wb.save(TMP_PATH)
    print(f"✅ 已保存 {TMP_PATH}(待切换)")
    print(f"Sheet 1: {ws1.max_row} 行(12 列)")
    print(f"Sheet 2: {ws2.max_row} 行")
    print(f"Sheet 3: {ws3.max_row} 行")
    print(f"Sheet 4: {ws4.max_row} 行(按 L1→L2→L3→L4→污染物 排序)")


if __name__ == "__main__":
    main()