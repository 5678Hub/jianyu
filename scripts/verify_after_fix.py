import openpyxl
wb = openpyxl.load_workbook('data/gb2762/_gb2762_2025_食品类别_v6.xlsx', data_only=True)
ws = wb['按食品类别查询']

print("=== Sheet 1 表 5 锡(4 行) ===")
for r in range(2, ws.max_row+1):
    if ws.cell(r, 8).value == '锡':
        cells = [str(ws.cell(r, c).value or '')[:18] for c in range(1, 13)]
        print('  | '.join(cells))

print("\n=== Sheet 1 表 9 苯并[a]芘(a 只在稻谷) ===")
for r in range(2, ws.max_row+1):
    if ws.cell(r, 8).value == '苯并[a]芘':
        cells = [str(ws.cell(r, c).value or '')[:18] for c in range(1, 13)]
        print('  | '.join(cells))

print("\n=== 包装饮用水 检验方法 ===")
for r in range(2, ws.max_row+1):
    food = ws.cell(r, 6).value or ''
    if '包装饮用水' in str(food) and '除外' not in str(food):
        cells = [str(ws.cell(r, c).value or '')[:18] for c in range(1, 13)]
        print('  | '.join(cells))

print("\n=== 各表条数 ===")
from collections import Counter
c = Counter()
for r in range(2, ws.max_row+1):
    p = ws.cell(r, 8).value
    if p:
        c[p] += 1
for p, n in sorted(c.items(), key=lambda x: -x[1]):
    print(f"  {p}: {n} 条")