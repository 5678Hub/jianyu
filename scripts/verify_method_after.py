import openpyxl
wb = openpyxl.load_workbook('data/gb2762/_gb2762_2025_食品类别_v6.xlsx', data_only=True)
ws = wb['按食品类别查询']

print("=== 包装饮用水 检验方法(应全部 GB 8538) ===")
for r in range(2, ws.max_row+1):
    food = ws.cell(r, 6).value or ''
    if '包装饮用水' in str(food) and '除外' not in str(food):
        cells = [str(ws.cell(r, c).value or '')[:18] for c in range(1, 13)]
        print('  | '.join(cells))

print("\n=== 表 5 锡 ===")
for r in range(2, ws.max_row+1):
    if ws.cell(r, 8).value == '锡':
        cells = [str(ws.cell(r, c).value or '')[:18] for c in range(1, 13)]
        print('  | '.join(cells))

print("\n=== 表 9 苯并[a]芘(看脚注列) ===")
for r in range(2, ws.max_row+1):
    if ws.cell(r, 8).value == '苯并[a]芘':
        l3 = ws.cell(r, 4).value
        foot = ws.cell(r, 11).value
        print(f"  L3={l3!r} 脚注={foot!r}")