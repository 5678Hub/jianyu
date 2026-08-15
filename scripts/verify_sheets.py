import openpyxl
wb = openpyxl.load_workbook('data/gb2762/_gb2762_2025_食品类别_v6.xlsx', data_only=True)
for name in wb.sheetnames:
    ws = wb[name]
    print(f"=== {name} ({ws.max_row} 行 × {ws.max_column} 列) ===")
    for r in range(1, min(ws.max_row + 1, 20)):
        cells = [str(ws.cell(r, c).value or '')[:30] for c in range(1, ws.max_column + 1)]
        print('  ' + ' | '.join(cells))
    print()